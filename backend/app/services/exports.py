"""Exportaciones: código .abap, Excel (casos de prueba) y PDF (spec, dump, inspector)."""
from __future__ import annotations
import io
from typing import Any


# ─── .abap ───────────────────────────────────────────────────────────────────
def code_to_abap(name: str, code: str) -> bytes:
    header = f"*&---------------------------------------------------------------------*\n*& {name}\n*& Generado por ABAP Factory AI\n*&---------------------------------------------------------------------*\n"
    return (header + (code or "")).encode("utf-8")


# ─── Excel ─────────────────────────────────────────────────────────────────--
def test_cases_to_xlsx(cases: list[dict], title: str = "Casos de prueba") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Casos"
    cols = ["case_id", "name", "preconditions", "input_data", "steps", "expected_result",
            "actual_result", "status", "evidence_required", "observations"]
    headers = ["ID", "Nombre", "Precondiciones", "Datos entrada", "Pasos", "Resultado esperado",
               "Resultado obtenido", "Estado", "Evidencia", "Observaciones"]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1E3A5F")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    for c in cases:
        ws.append([str(c.get(k, "") or "") for k in cols])
    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 24
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── PDF ───────────────────────────────────────────────────────────────────--
def _pdf_from_blocks(title: str, blocks: list[tuple[str, Any]]) -> bytes:
    """Genera un PDF simple. blocks = [(heading, content_str_or_list)]."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Preformatted
    from reportlab.lib import colors

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Title"], textColor=colors.HexColor("#1E3A5F"))
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=colors.HexColor("#2563EB"))
    body = styles["BodyText"]
    code = ParagraphStyle("code", parent=styles["Code"], fontSize=7.5, leading=9)

    flow = [Paragraph(title, h1), Spacer(1, 0.4 * cm)]
    for heading, content in blocks:
        if heading:
            flow.append(Paragraph(heading, h2))
        if isinstance(content, (list, tuple)):
            for item in content:
                flow.append(Paragraph(f"• {item}", body))
        elif content and ("\n" in str(content) or heading.lower().startswith("código")):
            flow.append(Preformatted(str(content)[:6000], code))
        else:
            flow.append(Paragraph(str(content or "—"), body))
        flow.append(Spacer(1, 0.3 * cm))
    doc.build(flow)
    return buf.getvalue()


def spec_to_pdf(spec: dict, title: str = "Especificación Técnica") -> bytes:
    blocks = [
        ("Descripción funcional", spec.get("functional_description")),
        ("Objetivo técnico", spec.get("technical_objective")),
        ("Supuestos", spec.get("assumptions")),
        ("Objetos SAP involucrados", spec.get("sap_objects")),
        ("Tablas estándar", spec.get("standard_tables")),
        ("BAPIs sugeridas", spec.get("suggested_bapis")),
        ("BAdIs / User-Exits", spec.get("badis_user_exits")),
        ("Riesgos", spec.get("risks")),
        ("Dependencias", spec.get("dependencies")),
        ("Performance", spec.get("performance_notes")),
        ("Seguridad", spec.get("security_notes")),
        ("Plan de transporte", spec.get("transport_plan")),
        ("Plan de rollback", spec.get("rollback_plan")),
    ]
    return _pdf_from_blocks(title, blocks)


def dump_to_pdf(d: dict, title: str = "Informe de Dump") -> bytes:
    blocks = [
        ("Tipo de dump", d.get("dump_type")),
        ("Severidad", d.get("severity")),
        ("Ubicación", f"Programa: {d.get('program')} | Include: {d.get('include')} | Línea: {d.get('line')} | Objeto: {d.get('sap_object')}"),
        ("Causa raíz", d.get("root_cause")),
        ("Solución", d.get("solution")),
        ("Código corregido", d.get("fixed_code")),
        ("Checklist de revisión", d.get("checklist")),
        ("Pruebas sugeridas", d.get("suggested_tests")),
    ]
    return _pdf_from_blocks(title, blocks)


def documentation_pdf(*, project: dict, requirement: dict | None, spec: dict | None,
                       artifacts: list[dict], inspections: list[dict],
                       test_suites: list[dict]) -> bytes:
    """Documento técnico completo de un desarrollo: portada + spec + código + calidad + pruebas."""
    blocks: list[tuple[str, Any]] = []
    blocks.append(("Proyecto", f"{project.get('name')} ({project.get('sap_version')}) · ambiente {project.get('environment')}"))
    if requirement:
        blocks.append(("Requerimiento", requirement.get("title")))
        blocks.append(("Descripción", requirement.get("description")))
    if spec:
        blocks.append(("— Especificación técnica —", ""))
        blocks.append(("Descripción funcional", spec.get("functional_description")))
        blocks.append(("Objetivo técnico", spec.get("technical_objective")))
        blocks.append(("Objetos SAP", spec.get("sap_objects")))
        blocks.append(("Riesgos", spec.get("risks")))
        blocks.append(("Plan de transporte", spec.get("transport_plan")))
        blocks.append(("Plan de rollback", spec.get("rollback_plan")))
    for a in artifacts:
        blocks.append((f"— Código: {a.get('name')} (v{a.get('version')}, {a.get('status')}) —", ""))
        blocks.append(("Código", a.get("code")))
        if a.get("explanation"):
            blocks.append(("Explicación", a.get("explanation")))
    for ins in inspections:
        blocks.append(("— Code Inspector —", f"Score {ins.get('score')}/100 · S/4: {ins.get('s4hana_compatible')}"))
        blocks.append(("Recomendación", ins.get("recommendation")))
    for ts in test_suites:
        blocks.append(("— Pruebas ABAP Unit —", f"Cobertura esperada: {ts.get('expected_coverage')}"))
        if ts.get("test_code"):
            blocks.append(("Código de test", ts.get("test_code")))
    return _pdf_from_blocks("Documentación Técnica del Desarrollo", blocks)


def migration_to_pdf(m: dict, title: str = "Informe de Migración ECC → S/4HANA") -> bytes:
    changes = [
        f"[{c.get('area')}] {c.get('reason')}\n  ANTES: {c.get('before')}\n  DESPUÉS: {c.get('after')}"
        for c in m.get("changes", [])
    ]
    si = [f"{s.get('item')} · {s.get('table')}: {s.get('note')}" for s in m.get("simplification_items", [])]
    blocks = [
        ("Destino", m.get("target")),
        ("Compatibilidad", m.get("compatibility")),
        ("Resumen", m.get("notes")),
        ("Código migrado", m.get("migrated_code")),
        ("Cambios aplicados", changes),
        ("Simplification items", si),
        ("Pasos manuales", m.get("manual_steps")),
    ]
    return _pdf_from_blocks(title, blocks)


def dev_document_pdf(d: dict, title: str | None = None) -> bytes:
    """Documento técnico: inventario de objetos + paso a paso (lo que el equipo construye)."""
    objs = [
        f"[{o.get('action', '').upper()}] {o.get('name')} ({o.get('type')}) — paquete {o.get('package') or '—'}\n"
        f"  {o.get('description') or ''}" + (f"\n  Depende de: {o.get('dependencies')}" if o.get('dependencies') else "")
        for o in d.get("objects", [])
    ]
    steps = [
        f"Paso {s.get('n')}: {s.get('title')}\n  {s.get('detail')}"
        + (f"\n  Objetos: {', '.join(s.get('objects', []))}" if s.get('objects') else "")
        for s in d.get("steps", [])
    ]
    blocks = [
        ("Resumen", d.get("summary")),
        (f"Inventario de objetos ({len(d.get('objects', []))})", objs),
        ("Construcción paso a paso", steps),
        ("Plan de transporte", d.get("transport_plan")),
        ("Plan de rollback", d.get("rollback_plan")),
    ]
    return _pdf_from_blocks(title or d.get("title") or "Documento Técnico del Desarrollo", blocks)


def inspection_to_pdf(ins: dict, title: str = "Informe Code Inspector / ATC") -> bytes:
    findings = [
        f"[{f.get('severity')}] {f.get('rule')} (línea {f.get('line')}): {f.get('message')} → {f.get('suggestion')}"
        for f in ins.get("findings", [])
    ]
    blocks = [
        ("Score técnico", f"{ins.get('score')}/100"),
        ("Compatibilidad S/4HANA", ins.get("s4hana_compatible")),
        ("Reglas incumplidas", ins.get("rules_violated")),
        ("Hallazgos", findings),
        ("Recomendación", ins.get("recommendation")),
        ("Código corregido", ins.get("corrected_code")),
    ]
    return _pdf_from_blocks(title, blocks)
